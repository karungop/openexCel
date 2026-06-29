"""
Speed benchmark: openexcel (C) vs openpyxl (pure Python)

Tests write and read performance at escalating row counts to make
the magnitude difference apparent.
"""

import datetime
import os
import sys
import tempfile
import time

import openexcel
import openpyxl

# ── data ──────────────────────────────────────────────────────────────────────

HEADERS = ["ID", "Name", "Score", "Active", "Date", "Notes"]

def make_row(i: int) -> list:
    return [
        i,
        f"User_{i}",
        round(i * 3.14159, 4),
        i % 2 == 0,
        datetime.date(2020, 1, 1) + datetime.timedelta(days=i % 365),
        f"This is a note for row {i} with some extra text to pad the string",
    ]


# ── timing helpers ─────────────────────────────────────────────────────────────

def timer(fn):
    t0 = time.perf_counter()
    fn()
    return time.perf_counter() - t0


def write_openexcel(path: str, n: int):
    wb = openexcel.Workbook()
    ws = wb.create_sheet("Sheet1")
    ws.append(HEADERS)
    for i in range(n):
        ws.append(make_row(i))
    wb.save(path)


def write_openpyxl(path: str, n: int):
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Sheet1")
    ws.append(HEADERS)
    for i in range(n):
        ws.append(make_row(i))
    wb.save(path)


def read_openexcel(path: str):
    wb = openexcel.load_workbook(path)
    ws = wb.active
    total = 0
    for row in ws.iter_rows():
        total += len(row)
    return total


def read_openpyxl(path: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    total = 0
    for row in ws.iter_rows():
        total += len(row)
    wb.close()
    return total


# ── benchmark ─────────────────────────────────────────────────────────────────

ROW_COUNTS = [1_000, 10_000, 50_000, 100_000, 250_000]

def bar(ratio: float, width: int = 30) -> str:
    filled = min(int(ratio), width)
    return "█" * filled + "░" * (width - filled)


def run():
    print()
    print("=" * 72)
    print("  openexcel (C)  vs  openpyxl (Python)  —  speed benchmark")
    print("=" * 72)

    tmpdir = tempfile.mkdtemp()
    path_oe = os.path.join(tmpdir, "oe.xlsx")
    path_px = os.path.join(tmpdir, "px.xlsx")

    # ── WRITE ──────────────────────────────────────────────────────────────────
    print()
    print("  WRITE  (append rows + save)")
    print(f"  {'Rows':>8}  {'openexcel':>10}  {'openpyxl':>10}  {'Speedup':>8}  Bar (openpyxl relative)")
    print("  " + "-" * 68)

    write_results = []
    for n in ROW_COUNTS:
        t_oe = timer(lambda: write_openexcel(path_oe, n))
        t_px = timer(lambda: write_openpyxl(path_px, n))
        ratio = t_px / t_oe
        write_results.append((n, t_oe, t_px, ratio))
        b = bar(ratio)
        print(f"  {n:>8,}  {t_oe:>9.3f}s  {t_px:>9.3f}s  {ratio:>6.1f}×   {b}")

    # ── READ ───────────────────────────────────────────────────────────────────
    print()
    print("  READ  (iter all rows, openexcel file vs openpyxl file)")
    print(f"  {'Rows':>8}  {'openexcel':>10}  {'openpyxl':>10}  {'Speedup':>8}  Bar (openpyxl relative)")
    print("  " + "-" * 68)

    for n, *_ in write_results:
        # re-write fresh files of exactly n rows so sizes match
        write_openexcel(path_oe, n)
        write_openpyxl(path_px, n)
        t_oe = timer(lambda: read_openexcel(path_oe))
        t_px = timer(lambda: read_openpyxl(path_px))
        ratio = t_px / t_oe
        b = bar(ratio)
        print(f"  {n:>8,}  {t_oe:>9.3f}s  {t_px:>9.3f}s  {ratio:>6.1f}×   {b}")

    # ── summary ────────────────────────────────────────────────────────────────
    print()
    print("=" * 72)
    n, t_oe, t_px, ratio = write_results[-1]
    print(f"  At {n:,} rows: openexcel wrote in {t_oe:.3f}s vs openpyxl {t_px:.3f}s  ({ratio:.1f}× faster)")
    print("=" * 72)
    print()

    # cleanup
    for p in (path_oe, path_px):
        try:
            os.unlink(p)
        except OSError:
            pass
    os.rmdir(tmpdir)


if __name__ == "__main__":
    run()
