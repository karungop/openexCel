"""
New feature benchmarks: openexcel (C) vs openpyxl (pure Python)

Benchmarks features beyond bulk row read/write:
  1. Cell object access (ws["A1"] and ws.cell())
  2. Number format set
  3. Number format roundtrip (save + reload)
  4. Formula write
  5. Formula roundtrip (save + reload)
  6. Merged cells (merge + save + reload)
  7. Column/row dimensions (set + save + reload)
  8. Named ranges (add + save + reload)
  9. Freeze panes (set + save + reload)
 10. Hyperlinks (50 cells, roundtrip)
 11. Data validation (10 rules, roundtrip)
 12. Page setup (roundtrip)
 13. Sheet protection (roundtrip)
"""

import os
import statistics
import tempfile
import time

import openexcel
import openpyxl
import openpyxl.workbook.defined_name as openpyxl_dn
from openpyxl.worksheet.datavalidation import DataValidation as OpenpyxlDV
from openpyxl.worksheet.protection import SheetProtection as OpenpyxlSP

ITERATIONS = 3


def median_ms(fn):
    """Run fn ITERATIONS times, return median in milliseconds."""
    times = []
    for _ in range(ITERATIONS):
        t0 = time.perf_counter()
        fn()
        times.append((time.perf_counter() - t0) * 1000)
    return statistics.median(times)


def fmt_ms(ms):
    if ms < 1:
        return f"{ms*1000:.1f} µs"
    elif ms < 1000:
        return f"{ms:.2f} ms"
    else:
        return f"{ms/1000:.3f} s"


# ── 1. Cell object access (10,000 reads) ─────────────────────────────────────

def bench_cell_access_openexcel():
    wb = openexcel.Workbook()
    ws = wb.create_sheet("Sheet1")
    # Populate 100×100 = 10,000 cells
    for r in range(1, 101):
        ws.append([r * c for c in range(1, 101)])

    def run():
        total = 0
        for r in range(1, 101):
            for c in range(1, 101):
                cell = ws.cell(row=r, column=c)
                total += 1
        return total

    return median_ms(run)


def bench_cell_access_openpyxl():
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(1, 101):
        ws.append([r * c for c in range(1, 101)])

    def run():
        total = 0
        for r in range(1, 101):
            for c in range(1, 101):
                cell = ws.cell(row=r, column=c)
                total += 1
        return total

    return median_ms(run)


# ── 2. Number format set (1,000 cells) ────────────────────────────────────────

def bench_numfmt_set_openexcel():
    wb = openexcel.Workbook()
    ws = wb.create_sheet("Sheet1")
    for i in range(1, 1001):
        ws.cell(row=i, column=1).value = i * 0.01

    def run():
        for i in range(1, 1001):
            ws.cell(row=i, column=1).number_format = "0.00%"

    return median_ms(run)


def bench_numfmt_set_openpyxl():
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(1, 1001):
        ws.cell(row=i, column=1).value = i * 0.01

    def run():
        for i in range(1, 1001):
            ws.cell(row=i, column=1).number_format = "0.00%"

    return median_ms(run)


# ── 3. Number format roundtrip ────────────────────────────────────────────────

def bench_numfmt_roundtrip_openexcel():
    path = tempfile.mktemp(suffix=".xlsx")

    def run():
        wb = openexcel.Workbook()
        ws = wb.create_sheet("Sheet1")
        for i in range(1, 1001):
            c = ws.cell(row=i, column=1)
            c.value = i * 0.01
            c.number_format = "0.00%"
        wb.save(path)
        wb2 = openexcel.load_workbook(path)
        ws2 = wb2.active
        for i in range(1, 1001):
            _ = ws2.cell(row=i, column=1).number_format

    t = median_ms(run)
    try:
        os.unlink(path)
    except OSError:
        pass
    return t


def bench_numfmt_roundtrip_openpyxl():
    path = tempfile.mktemp(suffix=".xlsx")

    def run():
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 1001):
            c = ws.cell(row=i, column=1)
            c.value = i * 0.01
            c.number_format = "0.00%"
        wb.save(path)
        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        for i in range(1, 1001):
            _ = ws2.cell(row=i, column=1).number_format

    t = median_ms(run)
    try:
        os.unlink(path)
    except OSError:
        pass
    return t


# ── 4. Formula write (1,000 cells) ───────────────────────────────────────────

def bench_formula_write_openexcel():
    wb = openexcel.Workbook()
    ws = wb.create_sheet("Sheet1")

    def run():
        for i in range(1, 1001):
            ws.cell(row=i, column=2).value = f"=SUM(A{i}:A{i})"

    return median_ms(run)


def bench_formula_write_openpyxl():
    wb = openpyxl.Workbook()
    ws = wb.active

    def run():
        for i in range(1, 1001):
            ws.cell(row=i, column=2).value = f"=SUM(A{i}:A{i})"

    return median_ms(run)


# ── 5. Formula roundtrip ─────────────────────────────────────────────────────

def bench_formula_roundtrip_openexcel():
    path = tempfile.mktemp(suffix=".xlsx")

    def run():
        wb = openexcel.Workbook()
        ws = wb.create_sheet("Sheet1")
        for i in range(1, 1001):
            ws.cell(row=i, column=1).value = i
            ws.cell(row=i, column=2).value = f"=SUM(A{i}:A{i})"
        wb.save(path)
        wb2 = openexcel.load_workbook(path)
        ws2 = wb2.active
        for i in range(1, 1001):
            _ = ws2.cell(row=i, column=2).value

    t = median_ms(run)
    try:
        os.unlink(path)
    except OSError:
        pass
    return t


def bench_formula_roundtrip_openpyxl():
    path = tempfile.mktemp(suffix=".xlsx")

    def run():
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 1001):
            ws.cell(row=i, column=1).value = i
            ws.cell(row=i, column=2).value = f"=SUM(A{i}:A{i})"
        wb.save(path)
        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        for i in range(1, 1001):
            _ = ws2.cell(row=i, column=2).value

    t = median_ms(run)
    try:
        os.unlink(path)
    except OSError:
        pass
    return t


# ── 6. Merged cells ──────────────────────────────────────────────────────────

def bench_merge_cells_openexcel():
    path = tempfile.mktemp(suffix=".xlsx")

    def run():
        wb = openexcel.Workbook()
        ws = wb.create_sheet("Sheet1")
        # 100 non-overlapping 1-row merges across columns A:C, rows 1..100
        for r in range(1, 101):
            start_col = "A"
            ws.merge_cells(f"A{r}:C{r}")
        wb.save(path)
        wb2 = openexcel.load_workbook(path)
        ws2 = wb2.active
        _ = ws2.merged_cells

    t = median_ms(run)
    try:
        os.unlink(path)
    except OSError:
        pass
    return t


def bench_merge_cells_openpyxl():
    path = tempfile.mktemp(suffix=".xlsx")

    def run():
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in range(1, 101):
            ws.merge_cells(f"A{r}:C{r}")
        wb.save(path)
        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        _ = ws2.merged_cells

    t = median_ms(run)
    try:
        os.unlink(path)
    except OSError:
        pass
    return t


# ── 7. Column/row dimensions ─────────────────────────────────────────────────

def bench_dimensions_openexcel():
    path = tempfile.mktemp(suffix=".xlsx")

    def run():
        wb = openexcel.Workbook()
        ws = wb.create_sheet("Sheet1")
        for col in range(1, 51):
            ws.set_column_width(col, 15 + col)
        for row in range(1, 51):
            ws.set_row_height(row, 20 + row)
        wb.save(path)
        wb2 = openexcel.load_workbook(path)
        _ = wb2.active

    t = median_ms(run)
    try:
        os.unlink(path)
    except OSError:
        pass
    return t


def bench_dimensions_openpyxl():
    path = tempfile.mktemp(suffix=".xlsx")

    col_letters = [
        openpyxl.utils.get_column_letter(c) for c in range(1, 51)
    ]

    def run():
        wb = openpyxl.Workbook()
        ws = wb.active
        for i, letter in enumerate(col_letters, 1):
            ws.column_dimensions[letter].width = 15 + i
        for row in range(1, 51):
            ws.row_dimensions[row].height = 20 + row
        wb.save(path)
        wb2 = openpyxl.load_workbook(path)
        _ = wb2.active

    t = median_ms(run)
    try:
        os.unlink(path)
    except OSError:
        pass
    return t


# ── 8. Named ranges ──────────────────────────────────────────────────────────

def bench_named_ranges_openexcel():
    path = tempfile.mktemp(suffix=".xlsx")

    def run():
        wb = openexcel.Workbook()
        wb.create_sheet("Sheet1")
        for i in range(1, 21):
            wb.add_defined_name(f"Range{i}", f"Sheet1!$A${i}:$C${i+5}")
        wb.save(path)
        wb2 = openexcel.load_workbook(path)
        _ = wb2.defined_names

    t = median_ms(run)
    try:
        os.unlink(path)
    except OSError:
        pass
    return t


def bench_named_ranges_openpyxl():
    path = tempfile.mktemp(suffix=".xlsx")

    def run():
        wb = openpyxl.Workbook()
        for i in range(1, 21):
            dn = openpyxl_dn.DefinedName(f"Range{i}", attr_text=f"Sheet1!$A${i}:$C${i+5}")
            wb.defined_names.add(dn)
        wb.save(path)
        wb2 = openpyxl.load_workbook(path)
        _ = wb2.defined_names

    t = median_ms(run)
    try:
        os.unlink(path)
    except OSError:
        pass
    return t


# ── 9. Freeze panes ──────────────────────────────────────────────────────────

def bench_freeze_panes_openexcel():
    path = tempfile.mktemp(suffix=".xlsx")

    def run():
        wb = openexcel.Workbook()
        ws = wb.create_sheet("Sheet1")
        for i in range(1, 101):
            ws.append([i, i * 2, i * 3])
        ws.freeze_panes = "B2"
        wb.save(path)
        wb2 = openexcel.load_workbook(path)
        ws2 = wb2.active
        _ = ws2.freeze_panes

    t = median_ms(run)
    try:
        os.unlink(path)
    except OSError:
        pass
    return t


def bench_freeze_panes_openpyxl():
    path = tempfile.mktemp(suffix=".xlsx")

    def run():
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 101):
            ws.append([i, i * 2, i * 3])
        ws.freeze_panes = "B2"
        wb.save(path)
        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        _ = ws2.freeze_panes

    t = median_ms(run)
    try:
        os.unlink(path)
    except OSError:
        pass
    return t


# ── main ──────────────────────────────────────────────────────────────────────

# ── 10. Hyperlinks (50 cells, roundtrip) ─────────────────────────────────────

def bench_hyperlinks_openexcel():
    def run():
        wb = openexcel.Workbook()
        ws = wb.create_sheet("Sheet1")
        for i in range(1, 51):
            ws.cell(row=i, column=1).value = f"Link {i}"
            ws.cell(row=i, column=1).hyperlink = f"https://example.com/{i}"
        tmp = tempfile.mktemp(suffix=".xlsx")
        wb.save(tmp)
        wb2 = openexcel.load_workbook(tmp)
        _ = [ws2.cell(row=i, column=1).hyperlink for i in range(1, 51)
             for ws2 in [wb2[0]]]
        os.unlink(tmp)
    return median_ms(run)

def bench_hyperlinks_openpyxl():
    def run():
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 51):
            ws.cell(row=i, column=1).value = f"Link {i}"
            ws.cell(row=i, column=1).hyperlink = f"https://example.com/{i}"
        tmp = tempfile.mktemp(suffix=".xlsx")
        wb.save(tmp)
        wb2 = openpyxl.load_workbook(tmp)
        ws2 = wb2.active
        _ = [ws2.cell(row=i, column=1).hyperlink for i in range(1, 51)]
        os.unlink(tmp)
    return median_ms(run)


# ── 11. Data validation (10 rules, roundtrip) ────────────────────────────────

def bench_data_validation_openexcel():
    def run():
        wb = openexcel.Workbook()
        ws = wb.create_sheet("Sheet1")
        for i in range(10):
            dv = openexcel.DataValidation(
                type="list", formula1='"Yes,No,Maybe"',
                sqref=f"A{i*10+1}:A{i*10+10}"
            )
            ws.add_data_validation(dv)
        tmp = tempfile.mktemp(suffix=".xlsx")
        wb.save(tmp)
        wb2 = openexcel.load_workbook(tmp)
        _ = wb2[0].data_validations
        os.unlink(tmp)
    return median_ms(run)

def bench_data_validation_openpyxl():
    def run():
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(10):
            dv = OpenpyxlDV(type="list", formula1='"Yes,No,Maybe"', showDropDown=False)
            dv.sqref = f"A{i*10+1}:A{i*10+10}"
            ws.add_data_validation(dv)
        tmp = tempfile.mktemp(suffix=".xlsx")
        wb.save(tmp)
        wb2 = openpyxl.load_workbook(tmp)
        _ = list(wb2.active.data_validations.dataValidation)
        os.unlink(tmp)
    return median_ms(run)


# ── 12. Page setup (roundtrip) ───────────────────────────────────────────────

def bench_page_setup_openexcel():
    def run():
        wb = openexcel.Workbook()
        ws = wb.create_sheet("Sheet1")
        ws.page_setup = openexcel.PageSetup(orientation="landscape", paper_size=9, scale=75)
        ws.page_margins = openexcel.PageMargins(left=0.5, right=0.5, top=1.0, bottom=1.0,
                                                 header=0.5, footer=0.5)
        tmp = tempfile.mktemp(suffix=".xlsx")
        wb.save(tmp)
        wb2 = openexcel.load_workbook(tmp)
        _ = wb2[0].page_setup.orientation
        _ = wb2[0].page_margins.left
        os.unlink(tmp)
    return median_ms(run)

def bench_page_setup_openpyxl():
    def run():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = 9
        ws.page_setup.scale = 75
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 1.0
        ws.page_margins.bottom = 1.0
        tmp = tempfile.mktemp(suffix=".xlsx")
        wb.save(tmp)
        wb2 = openpyxl.load_workbook(tmp)
        _ = wb2.active.page_setup.orientation
        _ = wb2.active.page_margins.left
        os.unlink(tmp)
    return median_ms(run)


# ── 13. Sheet protection (roundtrip) ─────────────────────────────────────────

def bench_sheet_protection_openexcel():
    def run():
        wb = openexcel.Workbook()
        ws = wb.create_sheet("Sheet1")
        ws.protection = openexcel.SheetProtection(sheet=True, format_cells=True)
        tmp = tempfile.mktemp(suffix=".xlsx")
        wb.save(tmp)
        wb2 = openexcel.load_workbook(tmp)
        _ = wb2[0].protection.sheet
        os.unlink(tmp)
    return median_ms(run)

def bench_sheet_protection_openpyxl():
    def run():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.protection.sheet = True
        ws.protection.formatCells = False
        tmp = tempfile.mktemp(suffix=".xlsx")
        wb.save(tmp)
        wb2 = openpyxl.load_workbook(tmp)
        _ = wb2.active.protection.sheet
        os.unlink(tmp)
    return median_ms(run)


BENCHMARKS = [
    (
        "Cell access (10k reads)",
        bench_cell_access_openexcel,
        bench_cell_access_openpyxl,
    ),
    (
        "Number format set (1k cells)",
        bench_numfmt_set_openexcel,
        bench_numfmt_set_openpyxl,
    ),
    (
        "Number format roundtrip (1k cells)",
        bench_numfmt_roundtrip_openexcel,
        bench_numfmt_roundtrip_openpyxl,
    ),
    (
        "Formula write (1k cells)",
        bench_formula_write_openexcel,
        bench_formula_write_openpyxl,
    ),
    (
        "Formula roundtrip (1k cells)",
        bench_formula_roundtrip_openexcel,
        bench_formula_roundtrip_openpyxl,
    ),
    (
        "Merged cells (100 ranges, roundtrip)",
        bench_merge_cells_openexcel,
        bench_merge_cells_openpyxl,
    ),
    (
        "Col/row dimensions (50+50, roundtrip)",
        bench_dimensions_openexcel,
        bench_dimensions_openpyxl,
    ),
    (
        "Named ranges (20, roundtrip)",
        bench_named_ranges_openexcel,
        bench_named_ranges_openpyxl,
    ),
    (
        "Freeze panes (roundtrip)",
        bench_freeze_panes_openexcel,
        bench_freeze_panes_openpyxl,
    ),
    (
        "Hyperlinks (50 cells, roundtrip)",
        bench_hyperlinks_openexcel,
        bench_hyperlinks_openpyxl,
    ),
    (
        "Data validation (10 rules, roundtrip)",
        bench_data_validation_openexcel,
        bench_data_validation_openpyxl,
    ),
    (
        "Page setup (roundtrip)",
        bench_page_setup_openexcel,
        bench_page_setup_openpyxl,
    ),
    (
        "Sheet protection (roundtrip)",
        bench_sheet_protection_openexcel,
        bench_sheet_protection_openpyxl,
    ),
]


def run():
    print()
    print("Running new-feature benchmarks (median of 3 iterations each)...")
    print()

    results = []
    for name, oe_fn, px_fn in BENCHMARKS:
        print(f"  {name}...", end="", flush=True)
        t_oe = oe_fn()
        t_px = px_fn()
        speedup = t_px / t_oe if t_oe > 0 else float("inf")
        results.append((name, t_oe, t_px, speedup))
        print(f" done ({fmt_ms(t_oe)} vs {fmt_ms(t_px)})")

    print()
    print("| Benchmark | openexcel | openpyxl | Speedup |")
    print("|---|---|---|---|")
    for name, t_oe, t_px, speedup in results:
        print(f"| {name} | {fmt_ms(t_oe)} | {fmt_ms(t_px)} | {speedup:.1f}× |")
    print()


if __name__ == "__main__":
    run()
